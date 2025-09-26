import os
from datetime import datetime
from openpyxl import Workbook
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
import requests
import urllib3
from urllib3.exceptions import InsecureRequestWarning

# Disable SSL warnings
os.environ['O365_DISABLE_SSL'] = 'true'
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Monkey-patch requests.Session.request to always use verify=False
old_request = requests.Session.request
def new_request(self, *args, **kwargs):
    kwargs['verify'] = False
    return old_request(self, *args, **kwargs)
requests.Session.request = new_request

# Set up file paths
folder = r"C:\Users\aritra.gautam\OneDrive - International SOS\Documents\Crowdstrike Falcon\Falcon Reports"

# Find the most recent CSV file that starts with "All_User_Accounts"
csv_files = [f for f in os.listdir(folder) if f.startswith("All_User_Accounts") and f.endswith(".csv")]
if not csv_files:
    raise FileNotFoundError("No CSV file starting with 'All_User_Accounts' found in the folder.")
csv_filename = max(csv_files, key=lambda f: os.path.getmtime(os.path.join(folder, f)))
csv_filepath = os.path.join(folder, csv_filename)

# Use a consistent Excel output name
now = datetime.now()
month_str = now.strftime('%b')
year_str = now.strftime('%Y')
xlsx_filename = f"Monthly Report ({month_str} {year_str}).xlsx"
xlsx_filepath = os.path.join(folder, xlsx_filename)

# SharePoint setup (use environment variables for credentials in production)
site_url = "https://internationalsosms.sharepoint.com/sites/GLBITSECCyberSecOps"
username = os.environ.get("SHAREPOINT_USERNAME", "cybersecurity.ops@internationalsos.com")
password = os.environ.get("SHAREPOINT_PASSWORD", "AutomationIsos223!@#")
target_folder_url = "/sites/GLBITSECCyberSecOps/Shared Documents/General/PowerBI/CrowdStrike"
archive_folder_url = "/sites/GLBITSECCyberSecOps/Shared Documents/General/PowerBI/CrowdStrike/Archive"

# 1. Read CSV and write to Excel
import csv

wb = Workbook()
ws = wb.active

with open(csv_filepath, 'r', encoding='utf-8-sig') as csvfile:
    reader = csv.reader(csvfile)
    for row in reader:
        ws.append(row)

# 2. Remove rows where Column D == "Unknown User"
rows_to_delete = [
    row[0].row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
    if len(row) >= 4 and row[3].value == "Unknown User"
]
for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

# 3. Fill blank fields in column G with domain from column D if D is not blank,
# and normalize column G if it ends with "orange2ca.gmessaging.net"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    col_d = row[3].value if len(row) > 3 else None
    col_g = row[6].value if len(row) > 6 else None
    # Fill blank G with domain from D
    if col_d and (col_g is None or str(col_g).strip() == "") and "@" in str(col_d):
        col_g = str(col_d).split("@", 1)[1]
        row[6].value = col_g
    # Normalize G if needed
    if col_g and str(col_g).lower().endswith("orange2ca.gmessaging.net"):
        row[6].value = "Orange2CA.gmessaging.net"

# Save the workbook after all Excel operations
wb.save(xlsx_filepath)
print(f"Excel report saved to {xlsx_filepath}")

# Ensure the file exists before upload
if not os.path.exists(xlsx_filepath):
    raise FileNotFoundError(f"Excel file not found after save: {xlsx_filepath}")

# 4. SharePoint operations
try:
    ctx = ClientContext(site_url).with_credentials(UserCredential(username, password))
    target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)
    files = target_folder.files
    ctx.load(files)
    ctx.execute_query()

    file_exists = False
    target_file = None
    file_list = []

    for f in files:
        if f.properties['Name'] == xlsx_filename:
            file_exists = True
            target_file = f
        # Parse creation date for all files
        time_created = f.properties.get('TimeCreated')
        if isinstance(time_created, str):
            try:
                f._created_dt = datetime.fromisoformat(time_created.replace('Z', '+00:00'))
            except Exception as ex:
                print(f"Failed to parse TimeCreated for {f.properties['Name']}: {ex}")
                continue
        elif isinstance(time_created, datetime):
            f._created_dt = time_created
        else:
            continue
        file_list.append(f)

    # Overwrite if file exists
    if file_exists and target_file is not None:
        try:
            target_file.delete_object().execute_query()
            print(f"Deleted existing {xlsx_filename} on SharePoint.")
        except Exception as del_exc:
            print(f"Failed to delete existing file: {del_exc}")
            raise

    # If file does not exist, archive the oldest file (if any and not the one being uploaded)
    elif not file_exists and file_list:
        file_list_no_new = [f for f in file_list if f.properties['Name'] != xlsx_filename]
        if file_list_no_new:
            oldest_file = min(file_list_no_new, key=lambda f: f._created_dt)
            oldest_file_name = oldest_file.properties['Name']
            dst_url = f"{archive_folder_url}"
            print(f"Moving file: {oldest_file_name} to {dst_url}")
            try:
                oldest_file.moveto(dst_url, 1).execute_query()
                print(f"Moved {oldest_file_name} to Archive.")
            except AttributeError:
                print("moveto() not available, falling back to copyto and delete.")
                oldest_file.copyto(dst_url, True).execute_query()
                oldest_file.delete_object().execute_query()
                print(f"Copied and deleted {oldest_file_name} to Archive.")

    # Upload new file
    with open(xlsx_filepath, 'rb') as content_file:
        target_folder.upload_file(xlsx_filename, content_file.read()).execute_query()
    print(f"Uploaded {xlsx_filename} to SharePoint.")

except Exception as e:
    import traceback
    print(f"SharePoint operation failed: {e}")
    traceback.print_exc()